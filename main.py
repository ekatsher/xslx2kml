# coding=utf-8
# This is a sample Python script.

# Press ⌃R to execute it or replace it with your code.
# Press Double ⇧ to search everywhere for classes, files, tool windows, actions, and settings.
from StdSuites import null

import requests

import openpyxl
from openpyxl import load_workbook
import xlsxwriter
import string
import lxml
import pykml as KML
from lxml import etree
from pykml.parser import Schema
from pykml.factory import KML_ElementMaker as KML
from pykml.factory import GX_ElementMaker as GX
import codecs
import sympy
from sympy import Polygon
import re

# в исходном файле в колонке 8 (считая с 0) - координаты точек, в колонке 9 - координаты полигонов

book = load_workbook("/Users/ekateshcherbakova/Documents/GermanyNoADM2.xlsx")  # ("/Users/ekateshcherbakova/Documents/KZ MATCH/Karagandinskaya oblast (1).xlsx")  # ("/Users/ekateshcherbakova/Documents/KZ MATCH/Ok/Karagandinskaya oblast  Ok.xlsx")
# book = openpyxl.open("/Users/ekateshcherbakova/Downloads/polygon.xlsx", read_only=True)

sheet = book.active

# create a KML file skeleton
doc = KML.kml(
    KML.Document(
        KML.Name("test"),
        KML.Style(
            KML.id("strokeColor:b51eff66_strokeWidth:8"),
            (KML.LineStyle(
                KML.Color("66ff1eb5"),
                KML.width(8)))
        )
    )
)

for row in range(2, sheet.max_row + 1):  # со второй строки, так как в первой заголовки
    # print(str(sheet[row][6].value))
    # name = str(sheet[row][6].value).decode('unicode')
    # преобразуем координаты полигонов
    new_poly_coord = str(sheet[row][8].value)
    # print(new_poly_coord.find('MULTI'))
    # stroka = 'POLIGON'
    # print(stroka.find('MULTI'))
    # если это не мультиполигон

    new_poly_coord = str(sheet[row][8].value).replace('POLYGON((', '')  # удаляем ненужную информацию 8 - для новых файлов, 10 - для исходных файлов
    new_poly_coord = new_poly_coord.replace('))', '')
    new_poly_coord = new_poly_coord.replace(' ', ';')  # заменяем пробелы на другой символ
    new_poly_coord = new_poly_coord.replace(',', ' ')  # заменяем пробелы запятыми (по формату необходимо чтобы координаты полигона были lat1,long1 lat2,long2 и т.д.)
    new_poly_coord = new_poly_coord.replace(';', ',')  # заменяем исходные пробелы запятыми, чтобы привести к правильному формату

    # преобразуем координаты точек

    new_point_coord = str(sheet[row][7].value).replace('POINT(',
                                                       '')  # удаляем ненужные символы 7 - для новых файлов, 9 - для исходных файлов
    new_point_coord = new_point_coord.replace(')', '')
    new_point_coord = new_point_coord.replace(' ', ',')

    # пока что избавляемся от мультиполигонов
    # print(new_poly_coord + ' ' + str(new_poly_coord.find('MULTI')))
    # print('MULTILINESTRING'.find('MULTI'))
    # print()

    if new_poly_coord.find('MULTI') >= 0:
        new_poly_coord = ''
    # ...и линий
    if new_poly_coord.find('LINESTRING') != -1:
        new_poly_coord = ''
    if new_poly_coord != 'None':
        # создаем полигоны
        pm2 = KML.Placemark(
            KML.description(sheet[row][0].value),
            KML.name(sheet[row][6].value),  # + ' ' + sheet[row][0].value),
            KML.Polygon(
                KML.outerBoundaryIs(
                    KML.LinearRing(
                        KML.coordinates(new_poly_coord)
                    )
                )
            )
        )
        doc.Document.append(pm2)
    # создаем точки
    pm1 = KML.Placemark(
        KML.description(sheet[row][0].value),
        KML.name(sheet[row][6].value),  # + ' ' + sheet[row][0].value),
        KML.Point(
            KML.coordinates(new_point_coord)
        )
    )

    # doc.Document.append(
    # KML.Folder(
    # pm1,
    # pm2))
    doc.Document.append(pm1)

#    print etree.tostring(pm1, pretty_print=True)
#    print etree.tostring(pm2, pretty_print=True)
print etree.tostring(doc, pretty_print=True)
my_file = open("/Users/ekateshcherbakova/Documents/GermanyNoADM2.kml", "w+")  # ("/Users/ekateshcherbakova/Documents/KZ MATCH/KML old/Karagandinskaya oblast old.kml", "w+")  # ("/Users/ekateshcherbakova/Documents/KZ MATCH/KML/Karagandinskaya oblast Ok.kml", "w+")
my_file.write(etree.tostring(doc, pretty_print=True))
my_file.close()

# print(row, sheet[row][9].value)

# print(new_cell)
# sheet[row][9].value = new_cell
# print(new_cell)
#    pieces = str(sheet[row][9].value).replace(",", " ").split()
# print(pieces[0])
# print(pieces)
# new_cell = ' '.join(','.join(pieces[i:i+2]) for i in xrange(0, len(pieces), 2))
#   for i in range(1, pieces.count(pieces)):
#       if i % 2 == 0:
#           pieces[i].join(",")
#   new_cell = pieces

#   print(row, new_cell)

# sheet[row][9].value = 99
# print(row, sheet[row][9].value)

# book.save("/Users/ekateshcherbakova/Downloads/savepolygon.xlsx")
